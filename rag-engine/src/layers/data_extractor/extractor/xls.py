from io import BytesIO
import uuid

from openpyxl import load_workbook
import xlrd

from src.layers.data_extractor.models import Line, Page, TablePage


def extract_data_excel(excel_bytes: bytes) -> tuple[list[Page], dict]:
    """
    Extract XLS / XLSX while preserving strict stream order:
    workbook â†’ sheets â†’ rows â†’ cells
    """

    pages: list[Page] = []
    page_number = 1

    metadata: dict[str, object] = {
        "_file_type": "excel",
        "_page_count": 0,
    }

    # -------- Try XLSX first --------
    try:
        wb = load_workbook(
            filename=BytesIO(excel_bytes),
            read_only=True,
            data_only=True,
        )

        for sheet in wb.worksheets:  # ðŸ”’ sheet order preserved
            page = _extract_xlsx_sheet(sheet, page_number)
            pages.append(page)
            page_number += 1

    # -------- Fallback to XLS --------
    except Exception:
        book = xlrd.open_workbook(file_contents=excel_bytes)

        for idx in range(book.nsheets):  # ðŸ”’ sheet order preserved
            sheet = book.sheet_by_index(idx)
            page = _extract_xls_sheet(sheet, page_number)
            pages.append(page)
            page_number += 1

    metadata["_page_count"] = len(pages)
    return pages, metadata


def _extract_xlsx_sheet(sheet, page_number: int) -> Page:
    lines: list[Line] = []
    tables: list[TablePage] = []

    y = 0.0

    rows = list(sheet.iter_rows(values_only=True))  # ðŸ”’ row order preserved

    if not rows:
        return Page(
            page_number=page_number,
            text="",
            lines=[],
            tables=[],
            images=[],
            width=None,
            height=None,
        )

    headers: list[str | None] = [str(c) if c is not None else None for c in rows[0]]

    table_rows: list[list[str | None]] = []

    for row in rows[1:]:  # ðŸ”’ row order preserved
        table_rows.append([
            str(cell) if cell is not None else None
            for cell in row  # ðŸ”’ cell order preserved
        ])

    tables.append(
        TablePage(
            id=str(uuid.uuid4()),
            bbox=(0, y, 0, y),
            data=[headers] + table_rows,
            top=y,
            x0=0,
            x1=800,
            bottom=y + 14 * (len(table_rows) + 1),
            page_number=page_number,
        )
    )

    # Optional text representation (useful for search)
    for row in rows:
        line_text = " | ".join(str(cell) if cell is not None else "" for cell in row)

        lines.append(
            Line(
                text=line_text,
                words=[],
                top=y,
                avg_size=12,
                is_bold=False,
                x0=0,
                x1=800,
                bottom=y + 12,
            )
        )
        y += 12

    return Page(
        page_number=page_number,
        text="\n".join(lin.text for lin in lines),
        lines=lines,
        tables=tables,
        images=[],
        width=None,
        height=None,
    )


def _extract_xls_sheet(sheet, page_number: int) -> Page:
    lines: list[Line] = []
    tables: list[TablePage] = []

    y = 0.0

    if sheet.nrows == 0:
        return Page(
            page_number=page_number,
            text="",
            lines=[],
            tables=[],
            images=[],
            width=None,
            height=None,
        )

    headers: list[str | None] = [
        str(sheet.cell_value(0, col)) if sheet.cell_value(0, col) != "" else None
        for col in range(sheet.ncols)
    ]

    table_rows: list[list[str | None]] = []

    for r in range(1, sheet.nrows):  # ðŸ”’ row order preserved
        table_rows.append([
            str(sheet.cell_value(r, c)) if sheet.cell_value(r, c) != "" else None
            for c in range(sheet.ncols)  # ðŸ”’ cell order preserved
        ])

    tables.append(
        TablePage(
            id=str(uuid.uuid4()),
            bbox=(0, y, 0, y),
            data=[headers] + table_rows,
            top=y,
            x0=0,
            x1=800,
            bottom=y + 14 * (len(table_rows) + 1),
            page_number=page_number,
        )
    )

    for r in range(sheet.nrows):
        row_text = " | ".join(
            str(sheet.cell_value(r, c)) if sheet.cell_value(r, c) != "" else ""
            for c in range(sheet.ncols)
        )

        lines.append(
            Line(
                text=row_text,
                words=[],
                top=y,
                avg_size=12,
                is_bold=False,
                x0=0,
                x1=800,
                bottom=y + 12,
            )
        )
        y += 12

    return Page(
        page_number=page_number,
        text="\n".join(lin.text for lin in lines),
        lines=lines,
        tables=tables,
        images=[],
        width=None,
        height=None,
    )
